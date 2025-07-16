import os, string, re, sys
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)


ircFile = open(sys.argv[1], 'r')
ircRaw  = []
for line in ircFile:
   llist = line.split()
   ircRaw.append(llist)
ircRawList = [_f for _f in ircRaw if _f]
points=len(ircRawList)

ircFile.close()


currentPath = os.getcwd()
fileName  = os.path.join(currentPath, 'pyfrag.csv')
coordFile = open(str(fileName), "w")
for Index, Molecule in enumerate(ircRawList):
   for atom in Molecule[1:-1]:
      coordFile.write(atom + ',')
   coordFile.write(Molecule[-1])
   coordFile.write('\n')
coordFile.close()



reload(sys)
sys.setdefaultencoding('utf8')

workbook = Workbook()
worksheet = workbook.active
with open(str(fileName), 'r') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        if r < 1:
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(',')):
                    cell = worksheet.cell(row=r+1, column=c+1)
                    cell.value = val
        elif r>1:
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(',')):
                    cell = worksheet.cell(row=r, column=c+1)
                    cell.value = float(val)
workbook.save('pyfrag.xlsx')



filename = os.path.join(currentPath, 'pyfrag.xlsx')

wb = load_workbook(filename)
ws = wb.active

chart = ScatterChart()

chart.style = 13
chart.x_axis.title = 'Bondlength'
chart.y_axis.title = 'Energy'

xvalues = Reference(ws, min_col=2, min_row=2, max_row=999)     #starting col including x xalue, starting data row and finishing data row

for i in [3,6,7]:                                          #cols including y values, noted final number is n+1 which is python range feature
    values = Reference(ws, min_col=i, min_row=1, max_row=999)  #rows of everything including first title row
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)


ws.add_chart(chart, "D10")



chart2 = ScatterChart()
chart2.style = 13
chart2.x_axis.title = 'Bondlength'
chart2.y_axis.title = 'Energy'

xvalues = Reference(ws, min_col=2, min_row=2, max_row=999)

for i in [4,5,6]:
    values = Reference(ws, min_col=i, min_row=1, max_row=999)
    series = Series(values, xvalues, title_from_data=True)
    chart2.series.append(series)


ws.add_chart(chart2, "D30")



wb.save(filename)

os.remove (fileName)
